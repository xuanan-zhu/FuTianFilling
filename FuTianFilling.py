import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext
from tkinter import ttk
import openpyxl
import re
from datetime import datetime
import os

# ================= 1. 配置区 =================

# 按照你要求的顺序定义列名
DEFAULT_HEADERS = [
    "团队", 
    "福田数量", 
    "序号", 
    "真实姓名", 
    "推荐人", 
    "居住地", 
    "职业", 
    "出身年月日", 
    "电话号码", 
    "现在生活事业家庭情况", 
    "想收获什么梦想", 
    "有无宗教信仰"
]

# ================= 2. 核心逻辑区 =================

def normalize_birth_date(value):
    """将各种格式的出生日期统一为：YYYY-MM-DD"""
    if not value:
        return ""
    nums = re.findall(r"\d+", value)
    if len(nums) >= 3:
        year, month, day = nums[:3]
        return f"{year}-{month.zfill(2)}-{day.zfill(2)}"
    return value

def extract_person_info(text):
    """解析文本提取信息"""
    text = text.replace("\r\n", "\n")

    field_alias = {
        "真实姓名": ["真实姓名", "姓名"],
        "推荐人": ["推荐人", "分享人"],
        "居住地": ["居住地", "地址"],
        "职业": ["职业"],
        "出身年月日": ["出身年月日", "出生年月日", "生日"],
        "电话号码": ["电话号码", "手机号码", "电话", "手机"],
        "现在生活事业家庭情况": ["现在生活事业家庭情况"],
        "想收获什么梦想": ["想收获什么梦想"],
        "有无宗教信仰": ["有无宗教信仰"]
    }

    reverse_map = {}
    for k, v in field_alias.items():
        for name in v:
            reverse_map[name] = k

    result = {k: "" for k in field_alias}
    current_field = None

    for raw_line in text.split("\n"):
        line = raw_line.strip()
        if not line:
            continue
        line = line.lstrip("0123456789. ")

        if "：" in line or ":" in line:
            key, val = line.replace(":", "：").split("：", 1)
            key = key.split("（")[0].split("(")[0].strip()
            key = key.replace(" ", "").replace("　", "")
            
            if key in reverse_map:
                current_field = reverse_map[key]
                if val.strip():
                    result[current_field] = val.strip()
                continue

        if current_field:
            if result[current_field]:
                result[current_field] += "\n" + line
            else:
                result[current_field] = line

    result["出身年月日"] = normalize_birth_date(result["出身年月日"])
    return result

def create_new_excel_file(file_path):
    """创建新的 Excel 文件并写入标准表头"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    # 写入表头
    ws.append(DEFAULT_HEADERS)
    
    # 设置一下列宽（美化）
    # A列(团队)到L列
    widths = {
        "A": 10, # 团队
        "B": 10, # 福田数量
        "C": 6,  # 序号
        "D": 12, # 姓名
        "E": 12, # 推荐人
        "H": 15, # 生日
        "I": 15, # 电话
    }
    
    for col_letter, width in widths.items():
         ws.column_dimensions[col_letter].width = width
    
    # 其他列默认宽一点
    for col in range(1, len(DEFAULT_HEADERS) + 1):
        letter = openpyxl.utils.get_column_letter(col)
        if letter not in widths:
            ws.column_dimensions[letter].width = 20
        
    wb.save(file_path)

def append_to_excel_safe(excel_path, text):
    """使用 openpyxl 追加数据，保留原有格式"""
    info = extract_person_info(text)

    try:
        wb = openpyxl.load_workbook(excel_path)
        sheet = wb.active
    except FileNotFoundError:
        raise Exception("找不到文件，请先创建或选择文件！")
    except Exception as e:
        raise Exception(f"打开 Excel 失败: {str(e)}")

    # 动态获取表头映射 {列名: 列索引}
    header_map = {}
    for col_idx, cell in enumerate(sheet[1], 1):
        if cell.value:
            header_map[str(cell.value).strip()] = col_idx

    if not header_map:
        raise Exception("Excel 文件似乎是空的（没有表头），请先检查或新建文件。")

    # 寻找最后一行
    next_row = sheet.max_row + 1
    
    # --- 1. 填入解析到的文本信息 ---
    for field, value in info.items():
        if field in header_map:
            col_index = header_map[field]
            sheet.cell(row=next_row, column=col_index).value = value

    # --- 2. 自动处理 '序号' 列 ---
    # 逻辑：如果表头里有“序号”这一列，我们就自动填入 (当前行号 - 1)
    if "序号" in header_map:
        seq_col = header_map["序号"]
        # 假设第一行是表头，那么第二行就是序号1
        seq_num = next_row - 1 
        sheet.cell(row=next_row, column=seq_col).value = seq_num

    # 注意："团队" 和 "福田数量" 因为文本里没有提取到，这里保持为空，你可以后续手动补
    
    try:
        wb.save(excel_path)
    except PermissionError:
        raise Exception("无法保存！请先关闭该 Excel 文件后再试。")
    
    return info

# ================= 3. GUI 界面区 =================

class AutoFillerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel 智能填表助手 v4.0 (定制版)")
        self.root.geometry("950x600")
        
        # 设置样式
        self.style = ttk.Style()
        self.style.theme_use('clam')
        self.style.configure("TButton", font=("微软雅黑", 9), padding=5)
        self.style.configure("Big.TButton", font=("微软雅黑", 11, "bold"))
        self.style.configure("TLabel", font=("微软雅黑", 10))
        self.style.configure("Header.TLabel", font=("微软雅黑", 12, "bold"))

        self.excel_path_var = tk.StringVar()
        
        self.setup_ui()

    def setup_ui(self):
        # --- 顶部：文件操作区 ---
        top_frame = ttk.LabelFrame(self.root, text="文件设置", padding=10)
        top_frame.pack(fill="x", padx=10, pady=5)

        ttk.Label(top_frame, text="当前 Excel:").pack(side="left")
        ttk.Entry(top_frame, textvariable=self.excel_path_var, width=50).pack(side="left", padx=5)
        
        # 按钮群
        ttk.Button(top_frame, text="📂 选择文件", command=self.choose_excel).pack(side="left", padx=2)
        ttk.Label(top_frame, text=" 或 ").pack(side="left")
        ttk.Button(top_frame, text="✨ 新建文件", command=self.create_excel).pack(side="left", padx=2)

        # --- 中部：主操作区 ---
        paned_window = ttk.PanedWindow(self.root, orient="horizontal")
        paned_window.pack(fill="both", expand=True, padx=10, pady=5)

        # === 左侧：输入区 ===
        left_frame = ttk.Frame(paned_window)
        paned_window.add(left_frame, weight=6)

        ttk.Label(left_frame, text="在此粘贴个人信息文本:", style="Header.TLabel").pack(anchor="w", pady=(0, 5))
        
        # 文本框
        self.text_input = scrolledtext.ScrolledText(left_frame, width=40, height=20, font=("Consolas", 10))
        self.text_input.pack(fill="both", expand=True)

        # 左侧底部按钮
        btn_frame = ttk.Frame(left_frame)
        btn_frame.pack(fill="x", pady=10)
        
        self.btn_run = ttk.Button(btn_frame, text="⚡ 立即追加到 Excel", style="Big.TButton", command=self.run_append)
        self.btn_run.pack(fill="x", ipady=5)
        
        ttk.Button(btn_frame, text="清空输入框", command=lambda: self.text_input.delete("1.0", tk.END)).pack(fill="x", pady=5)

        # === 右侧：历史记录区 ===
        right_frame = ttk.Frame(paned_window)
        paned_window.add(right_frame, weight=4)

        ttk.Label(right_frame, text="本次操作历史:", style="Header.TLabel").pack(anchor="w", pady=(0, 5), padx=5)
        
        # 表格 (Treeview)
        cols = ("name", "phone", "job", "time")
        self.tree = ttk.Treeview(right_frame, columns=cols, show="headings", height=20)
        
        self.tree.heading("name", text="姓名")
        self.tree.heading("phone", text="电话")
        self.tree.heading("job", text="职业")
        self.tree.heading("time", text="时间")
        
        self.tree.column("name", width=70)
        self.tree.column("phone", width=90)
        self.tree.column("job", width=70)
        self.tree.column("time", width=70)

        scrollbar = ttk.Scrollbar(right_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        
        tree_frame = ttk.Frame(right_frame)
        tree_frame.pack(fill="both", expand=True, padx=5)
        self.tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # 右侧底部：清空历史按钮
        ttk.Button(right_frame, text="🗑️ 清空历史记录", command=self.clear_history).pack(fill="x", padx=5, pady=10)

    # --- 功能函数 ---
    
    def choose_excel(self):
        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if path:
            self.excel_path_var.set(path)

    def create_excel(self):
        # 弹出保存对话框
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile="团队统计表.xlsx"
        )
        if path:
            try:
                create_new_excel_file(path)
                self.excel_path_var.set(path)
                messagebox.showinfo("成功", "新文件创建成功！\n表头已按指定格式生成。")
            except Exception as e:
                messagebox.showerror("创建失败", str(e))

    def run_append(self):
        excel_path = self.excel_path_var.get()
        text = self.text_input.get("1.0", tk.END).strip()

        if not excel_path:
            messagebox.showwarning("提示", "请先 [选择文件] 或 [新建文件]！")
            return
        if not text:
            messagebox.showwarning("提示", "文本框是空的！")
            return
        
        if not os.path.exists(excel_path):
             messagebox.showerror("错误", "指定的文件不存在，请重新选择或新建！")
             return

        try:
            extracted_info = append_to_excel_safe(excel_path, text)
            self.add_to_history(extracted_info)
            messagebox.showinfo("成功", f"已添加：{extracted_info.get('真实姓名', '未知')}")
            self.text_input.delete("1.0", tk.END)
            
        except Exception as e:
            messagebox.showerror("处理失败", str(e))

    def add_to_history(self, info):
        """添加到右侧列表"""
        name = info.get("真实姓名", "-")
        phone = info.get("电话号码", "-")
        job = info.get("职业", "-")
        current_time = datetime.now().strftime("%H:%M:%S")
        self.tree.insert("", 0, values=(name, phone, job, current_time))

    def clear_history(self):
        """清空右侧 Treeview 列表"""
        if not self.tree.get_children():
            return
        
        items = self.tree.get_children()
        for item in items:
            self.tree.delete(item)

if __name__ == "__main__":
    root = tk.Tk()
    app = AutoFillerApp(root)
    root.mainloop()