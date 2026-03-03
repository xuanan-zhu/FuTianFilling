import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext
from tkinter import ttk
import openpyxl
import re
from datetime import datetime
import os

# ================= 1. 配置区 =================

# --- 模式1：福田统计表头 ---
HEADERS_FUTIAN = [
    "团队", "福田数量", "序号", "真实姓名", "推荐人", 
    "居住地", "职业", "出身年月日", "电话号码", 
    "现在生活事业家庭情况", "想收获什么梦想", "有无宗教信仰"
]

# --- 模式2：爱心流动表头 ---
HEADERS_LOVE = [
    "被流动人", "类型", "份数", "日期", "流动人", 
    "回流人", "归属", "源头", "备注"
]

# ================= 2. 核心逻辑区 =================

def normalize_date(value):
    """通用日期清洗"""
    if not value: return ""
    value = value.replace("/", "-").replace(".", "-").replace("年", "-").replace("月", "-").replace("日", "")
    nums = re.findall(r"\d+", value)
    if len(nums) >= 3:
        year, month, day = nums[:3]
        if len(year) == 2: year = "20" + year
        return f"{year}-{month.zfill(2)}-{day.zfill(2)}"
    return value

def extract_futian_info(text):
    """【模式1】福田解析"""
    field_alias = {
        "真实姓名": ["真实姓名", "姓名"],
        "推荐人": ["推荐人", "分享人"],
        "居住地": ["居住地", "地址"],
        "职业": ["职业"],
        "出身年月日": ["出身年月日", "出生年月日", "生日"],
        "电话号码": ["电话号码", "手机号码", "电话", "手机"],
        "现在生活事业家庭情况": ["现在生活事业家庭情况"],
        "想收获什么梦想": ["想收获什么梦想"],
        "有无宗教信仰": ["有无宗教信仰"]
    }
    result = parse_key_value_text(text, field_alias)
    result["出身年月日"] = normalize_date(result.get("出身年月日", ""))
    return result

def extract_love_info(text):
    """【模式2】爱心解析 (修正版)"""
    field_alias = {
        "被流动人": ["被流动人", "被流动学员"],
        "原始类型": ["类型"], # 临时字段，用于后续拆分
        "日期": ["日期", "时间"],
        "流动人": ["流动人"],
        "回流人": ["回流人"],
        "归属": ["归属"],
        "源头": ["源头"],
        "备注": ["备注"]
    }
    result = parse_key_value_text(text, field_alias)
    
    # --- 修正：拆分类型和份数，并去除'份'字 ---
    raw_type = result.get("原始类型", "")
    result["类型"] = raw_type # 默认值
    result["份数"] = ""       # 默认值
    
    if raw_type:
        # 匹配中文括号或英文括号
        match = re.match(r"(.*?)[\(（](.*?)[\)）]", raw_type)
        if match:
            # 提取类型名称（如：爱心）
            result["类型"] = match.group(1).strip()
            
            # 提取括号内的内容（如：1份），并去掉“份”字
            raw_amount = match.group(2).strip()
            result["份数"] = raw_amount.replace("份", "").strip()
            
    result["日期"] = normalize_date(result.get("日期", ""))
    return result

def extract_custom_info(text):
    """【模式3】自定义解析"""
    text = text.replace("\r\n", "\n")
    result = {}
    current_key = None
    
    for raw_line in text.split("\n"):
        line = raw_line.strip()
        if not line: continue
        if line.startswith("【") and line.endswith("】"): continue

        if "：" in line or ":" in line:
            split_char = "：" if "：" in line else ":"
            parts = line.split(split_char, 1)
            key = parts[0].strip()
            val = parts[1].strip() if len(parts) > 1 else ""
            
            key_clean = key.split("（")[0].split("(")[0].strip().replace(" ", "").replace("　", "")
            current_key = key_clean
            result[current_key] = val
        elif current_key:
            result[current_key] += " " + line
    return result

def parse_key_value_text(text, field_alias):
    """通用解析内核"""
    reverse_map = {}
    for k, v in field_alias.items():
        for name in v:
            reverse_map[name] = k

    result = {k: "" for k in field_alias}
    current_field = None

    for raw_line in text.replace("\r\n", "\n").split("\n"):
        line = raw_line.strip()
        if not line: continue
        line = line.replace("【🔔流动明细表】", "").strip()
        if not line: continue

        if "：" in line or ":" in line:
            split_char = "：" if "：" in line else ":"
            parts = line.split(split_char, 1)
            key = parts[0].split("（")[0].split("(")[0].strip().replace(" ", "")
            val = parts[1].strip() if len(parts) > 1 else ""
            
            if key in reverse_map:
                current_field = reverse_map[key]
                if val: result[current_field] = val
                continue
        
        if current_field:
            if result[current_field]:
                result[current_field] += " " + line
            else:
                result[current_field] = line
    return result

def create_new_excel_file(file_path, headers):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
    wb.save(file_path)

def append_to_excel_safe(excel_path, text, mode):
    if mode == 1:
        info = extract_futian_info(text)
    elif mode == 2:
        info = extract_love_info(text)
    else:
        info = extract_custom_info(text)

    try:
        wb = openpyxl.load_workbook(excel_path)
        sheet = wb.active
    except Exception as e:
        raise Exception(f"打开 Excel 失败: {str(e)}")

    header_map = {}
    for col_idx, cell in enumerate(sheet[1], 1):
        if cell.value:
            header_map[str(cell.value).strip()] = col_idx

    if not header_map:
        raise Exception("Excel 文件没有表头，无法匹配数据。")

    next_row = sheet.max_row + 1
    
    for field, value in info.items():
        if field in header_map:
            sheet.cell(row=next_row, column=header_map[field]).value = value
        else:
            if mode == 3: # 简单模糊匹配
                for h_name, h_idx in header_map.items():
                     if field == h_name: # 精确匹配优先
                         sheet.cell(row=next_row, column=h_idx).value = value

    if "序号" in header_map:
        sheet.cell(row=next_row, column=header_map["序号"]).value = next_row - 1
    
    try:
        wb.save(excel_path)
    except PermissionError:
        raise Exception("无法保存！请先关闭该 Excel 文件后再试。")
    
    return info

# ================= 3. GUI 界面区 =================

class AutoFillerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel 智能填表助手 v7.0 (优化版)")
        self.root.geometry("1000x700") # 稍微加宽一点以容纳更多列
        
        self.excel_path_var = tk.StringVar()
        self.mode_var = tk.IntVar(value=1)
        self.custom_headers_var = tk.StringVar()
        
        self.style = ttk.Style()
        self.style.theme_use('clam')
        self.style.configure("Header.TLabel", font=("微软雅黑", 12, "bold"))
        self.style.configure("Big.TButton", font=("微软雅黑", 11, "bold"))
        
        self.setup_ui()

    def setup_ui(self):
        # --- 模式选择 ---
        mode_frame = ttk.LabelFrame(self.root, text="第一步：选择填表模式", padding=10)
        mode_frame.pack(fill="x", padx=10, pady=5)
        
        ttk.Radiobutton(mode_frame, text="类型一：福田统计", variable=self.mode_var, value=1, command=self.on_mode_change).grid(row=0, column=0, padx=20, sticky="w")
        ttk.Radiobutton(mode_frame, text="类型二：爱心流动", variable=self.mode_var, value=2, command=self.on_mode_change).grid(row=0, column=1, padx=20, sticky="w")
        ttk.Radiobutton(mode_frame, text="类型三：自定义", variable=self.mode_var, value=3, command=self.on_mode_change).grid(row=0, column=2, padx=20, sticky="w")
        
        self.custom_frame = ttk.Frame(mode_frame)
        self.custom_frame.grid(row=1, column=0, columnspan=3, sticky="we", pady=(10,0))
        ttk.Label(self.custom_frame, text="新建列名 (空格隔开):", foreground="blue").pack(side="left")
        ttk.Entry(self.custom_frame, textvariable=self.custom_headers_var, width=60).pack(side="left", padx=5)
        self.custom_frame.grid_remove()

        # --- 文件设置 ---
        file_frame = ttk.LabelFrame(self.root, text="第二步：文件设置", padding=10)
        file_frame.pack(fill="x", padx=10, pady=5)
        ttk.Label(file_frame, text="Excel路径:").pack(side="left")
        ttk.Entry(file_frame, textvariable=self.excel_path_var, width=50).pack(side="left", padx=5)
        ttk.Button(file_frame, text="📂 选择", command=self.choose_excel).pack(side="left")
        ttk.Label(file_frame, text=" | ").pack(side="left")
        ttk.Button(file_frame, text="✨ 新建", command=self.create_excel).pack(side="left")

        # --- 主操作区 ---
        paned = ttk.PanedWindow(self.root, orient="horizontal")
        paned.pack(fill="both", expand=True, padx=10, pady=5)

        # 左侧
        left_frame = ttk.Frame(paned)
        paned.add(left_frame, weight=5)
        ttk.Label(left_frame, text="粘贴文本:", style="Header.TLabel").pack(anchor="w")
        self.text_input = scrolledtext.ScrolledText(left_frame, width=40, height=20, font=("Consolas", 10))
        self.text_input.pack(fill="both", expand=True)
        
        btn_frame = ttk.Frame(left_frame)
        btn_frame.pack(fill="x", pady=10)
        ttk.Button(btn_frame, text="⚡ 写入 Excel", style="Big.TButton", command=self.run_append).pack(fill="x", ipady=5)
        ttk.Button(btn_frame, text="清空输入", command=lambda: self.text_input.delete("1.0", tk.END)).pack(fill="x", pady=5)

        # 右侧 (历史)
        right_frame = ttk.Frame(paned)
        paned.add(right_frame, weight=5)
        ttk.Label(right_frame, text="操作历史:", style="Header.TLabel").pack(anchor="w", padx=5)
        
        # 增加一列 c4，用于显示份数
        self.cols = ("c1", "c2", "c3", "c4", "time")
        self.tree = ttk.Treeview(right_frame, columns=self.cols, show="headings", height=20)
        
        # 设置列宽
        self.tree.column("c1", width=80)
        self.tree.column("c2", width=80)
        self.tree.column("c3", width=60)
        self.tree.column("c4", width=80)
        self.tree.column("time", width=70)
        
        self.update_history_header()
        
        scrollbar = ttk.Scrollbar(right_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        self.tree.pack(side="left", fill="both", expand=True, padx=5)
        scrollbar.pack(side="right", fill="y")
        
        ttk.Button(right_frame, text="🗑️ 清空历史", command=self.clear_history).pack(fill="x", padx=5, pady=10)

    # --- 逻辑 ---
    def on_mode_change(self):
        if self.mode_var.get() == 3: self.custom_frame.grid()
        else: self.custom_frame.grid_remove()
        self.update_history_header()

    def update_history_header(self):
        """根据模式动态调整表头显示"""
        mode = self.mode_var.get()
        if mode == 1:
            self.tree.heading("c1", text="姓名")
            self.tree.heading("c2", text="居住地")
            self.tree.heading("c3", text="电话")
            self.tree.heading("c4", text="职业")
        elif mode == 2:
            self.tree.heading("c1", text="被流动人")
            self.tree.heading("c2", text="类型")
            self.tree.heading("c3", text="份数") # 新增
            self.tree.heading("c4", text="流动人")
        else:
            self.tree.heading("c1", text="列1")
            self.tree.heading("c2", text="列2")
            self.tree.heading("c3", text="列3")
            self.tree.heading("c4", text="列4")
        self.tree.heading("time", text="时间")

    def choose_excel(self):
        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if path: self.excel_path_var.set(path)

    def create_excel(self):
        mode = self.mode_var.get()
        if mode == 1:
            headers, name = HEADERS_FUTIAN, "福田统计表.xlsx"
        elif mode == 2:
            headers, name = HEADERS_LOVE, "爱心流动表.xlsx"
        else:
            raw = self.custom_headers_var.get().strip()
            if not raw: return messagebox.showwarning("提示", "请输入列名！")
            headers = [h for h in re.split(r'[，, \s]+', raw) if h]
            name = "自定义表.xlsx"

        path = filedialog.asksaveasfilename(defaultextension=".xlsx", initialfile=name)
        if path:
            create_new_excel_file(path, headers)
            self.excel_path_var.set(path)
            messagebox.showinfo("成功", "文件创建成功！")

    def run_append(self):
        path, text = self.excel_path_var.get(), self.text_input.get("1.0", tk.END).strip()
        mode = self.mode_var.get()
        if not path or not os.path.exists(path): return messagebox.showerror("错误", "文件不存在！")
        if not text: return

        try:
            info = append_to_excel_safe(path, text, mode)
            self.add_to_history(info, mode)
            
            name = "未知"
            if mode == 1: name = info.get("真实姓名")
            elif mode == 2: name = info.get("被流动人")
            else: name = list(info.values())[0] if info else "记录"
            
            messagebox.showinfo("成功", f"已添加：{name}")
            self.text_input.delete("1.0", tk.END)
        except Exception as e:
            messagebox.showerror("错误", str(e))

    def add_to_history(self, info, mode):
        t = datetime.now().strftime("%H:%M:%S")
        vals = ["-", "-", "-", "-", t] # 默认为5个占位
        
        if mode == 1:
            vals[0] = info.get("真实姓名", "-")
            vals[1] = info.get("居住地", "-")
            vals[2] = info.get("电话号码", "-")
            vals[3] = info.get("职业", "-")
        elif mode == 2:
            vals[0] = info.get("被流动人", "-")
            vals[1] = info.get("类型", "-")
            vals[2] = info.get("份数", "-") # 对应界面上的“份数”列
            vals[3] = info.get("流动人", "-")
        else:
            v = list(info.values())
            for i in range(min(4, len(v))):
                vals[i] = v[i]
            
        self.tree.insert("", 0, values=vals)

    def clear_history(self):
        for item in self.tree.get_children(): self.tree.delete(item)

if __name__ == "__main__":
    root = tk.Tk()
    app = AutoFillerApp(root)
    root.mainloop()