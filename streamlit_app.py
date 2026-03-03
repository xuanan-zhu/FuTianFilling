import streamlit as st
import openpyxl
from openpyxl.styles import Alignment
import re
import io
import pandas as pd
from datetime import datetime

# ================= 1. 配置与解析逻辑区 =================

# --- 表头定义 ---
HEADERS_FUTIAN = [
    "团队", "福田数量", "序号", "真实姓名", "推荐人", 
    "居住地", "职业", "出身年月日", "电话号码", 
    "现在生活事业家庭情况", "想收获什么梦想", "有无宗教信仰"
]

HEADERS_LOVE = [
    "被流动人", "类型", "份数", "日期", "流动人", 
    "回流人", "归属", "源头", "备注"
]

# --- 解析函数 ---

def normalize_date(value):
    """通用日期清洗"""
    if not value: return ""
    value = str(value).replace("/", "-").replace(".", "-").replace("年", "-").replace("月", "-").replace("日", "")
    nums = re.findall(r"\d+", value)
    if len(nums) >= 3:
        year, month, day = nums[:3]
        if len(year) == 2: year = "20" + year
        return f"{year}-{month.zfill(2)}-{day.zfill(2)}"
    return value

def extract_info_by_mode(text, mode):
    """根据模式分发解析逻辑"""
    text = text.replace("\r\n", "\n")
    
    if mode == "福田统计":
        return extract_futian(text)
    elif mode == "爱心流动":
        return extract_love(text)
    else:
        return extract_custom(text)

def extract_futian(text):
    field_alias = {
        "真实姓名": ["真实姓名", "姓名"],
        "推荐人": ["推荐人", "分享人"],
        "居住地": ["居住地", "地址"],
        "职业": ["职业"],
        "出身年月日": ["出身年月日", "出生年月日", "生日"],
        "电话号码": ["电话号码", "手机号码", "电话", "手机"],
        "现在生活事业家庭情况": ["现在生活事业家庭情况"],
        "想收获什么梦想": ["想收获什么梦想"],
        "有无宗教信仰": ["有无宗教信仰"]
    }
    result = parse_key_value(text, field_alias)
    result["出身年月日"] = normalize_date(result.get("出身年月日", ""))
    return result

def extract_love(text):
    field_alias = {
        "被流动人": ["被流动人", "被流动学员"],
        "原始类型": ["类型"], 
        "日期": ["日期", "时间"],
        "流动人": ["流动人"],
        "回流人": ["回流人"],
        "归属": ["归属"],
        "源头": ["源头"],
        "备注": ["备注"]
    }
    result = parse_key_value(text, field_alias)
    
    # 拆分类型与份数
    raw_type = result.get("原始类型", "")
    result["类型"] = raw_type 
    result["份数"] = ""       
    
    if raw_type:
        # 匹配 "爱心(1份)" -> group1=爱心, group2=1份
        match = re.match(r"(.*?)[\(（](.*?)[\)）]", raw_type)
        if match:
            result["类型"] = match.group(1).strip()
            # 去掉 '份' 字
            result["份数"] = match.group(2).replace("份", "").strip()
            
    result["日期"] = normalize_date(result.get("日期", ""))
    return result

def extract_custom(text):
    result = {}
    current_key = None
    for raw_line in text.split("\n"):
        line = raw_line.strip()
        if not line: continue
        if line.startswith("【") and line.endswith("】"): continue

        if "：" in line or ":" in line:
            split_char = "：" if "：" in line else ":"
            parts = line.split(split_char, 1)
            key = parts[0].strip()
            val = parts[1].strip() if len(parts) > 1 else ""
            key_clean = key.split("（")[0].split("(")[0].strip().replace(" ", "")
            current_key = key_clean
            result[current_key] = val
        elif current_key:
            result[current_key] += " " + line
    return result

def parse_key_value(text, field_alias):
    reverse_map = {}
    for k, v in field_alias.items():
        for name in v: reverse_map[name] = k
    result = {k: "" for k in field_alias}
    current_field = None
    for raw_line in text.split("\n"):
        line = raw_line.strip()
        if not line: continue
        line = line.replace("【🔔流动明细表】", "").strip() # 清理标题
        if "：" in line or ":" in line:
            parts = line.replace(":", "：").split("：", 1)
            key = parts[0].split("（")[0].split("(")[0].strip().replace(" ", "")
            val = parts[1].strip() if len(parts) > 1 else ""
            if key in reverse_map:
                current_field = reverse_map[key]
                if val: result[current_field] = val
                continue
        if current_field and result[current_field]:
            result[current_field] += " " + line
        elif current_field:
            result[current_field] = line
    return result

# ================= 2. Excel 核心操作 (Session State版) =================

def create_blank_workbook(mode, custom_headers_str=""):
    """创建空白 Workbook 并设置表头"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    headers = []
    if mode == "福田统计":
        headers = HEADERS_FUTIAN
    elif mode == "爱心流动":
        headers = HEADERS_LOVE
    else:
        # 自定义模式
        headers = [h for h in re.split(r'[，, \s]+', custom_headers_str) if h]
        if not headers: headers = ["列名1", "列名2", "列名3"] # 默认保底

    ws.append(headers)
    
    # 设置列宽
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
        
    return wb

def append_data_to_workbook(wb, info_dict, mode):
    """将字典数据追加到 Workbook"""
    sheet = wb.active
    header_map = {}
    for col_idx, cell in enumerate(sheet[1], 1):
        if cell.value: header_map[str(cell.value).strip()] = col_idx
            
    if not header_map: return False, "表格没有表头，无法识别列名"

    next_row = sheet.max_row + 1
    
    for field, value in info_dict.items():
        # 1. 精确匹配
        if field in header_map:
            cell = sheet.cell(row=next_row, column=header_map[field])
            cell.value = value
            cell.alignment = Alignment(wrap_text=True) # 自动换行
        # 2. 模糊匹配 (仅自定义模式)
        elif mode == "自定义":
             for h_name, h_idx in header_map.items():
                if field == h_name:
                    sheet.cell(row=next_row, column=h_idx).value = value

    # 序号自动生成
    if "序号" in header_map:
        sheet.cell(row=next_row, column=header_map["序号"]).value = next_row - 1

    # 返回成功消息的关键字段
    key_name = "未知"
    if mode == "福田统计": key_name = info_dict.get("真实姓名", "未知")
    elif mode == "爱心流动": key_name = info_dict.get("被流动人", "未知")
    else: key_name = list(info_dict.values())[0] if info_dict else "数据"

    return True, f"成功添加：{key_name}"

def to_excel_bytes(wb):
    """将 Workbook 转换为二进制流供下载"""
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# ================= 3. Streamlit 界面交互 =================

st.set_page_config(page_title="Excel 智能填表助手 Pro", page_icon="📝", layout="wide")

# --- 初始化 Session State ---
if 'workbook' not in st.session_state: st.session_state.workbook = None
if 'file_name' not in st.session_state: st.session_state.file_name = "导出数据.xlsx"
if 'last_loaded_key' not in st.session_state: st.session_state.last_loaded_key = None
if 'status_msg' not in st.session_state: st.session_state.status_msg = None
# 默认模式
if 'current_mode' not in st.session_state: st.session_state.current_mode = "福田统计"

# --- 回调函数：处理提交 ---
def submit_data():
    text = st.session_state.user_input
    mode = st.session_state.current_mode
    
    if not text.strip():
        st.session_state.status_msg = ("warning", "⚠️ 内容不能为空！")
        return

    if st.session_state.workbook is None:
        st.session_state.status_msg = ("error", "❌ 请先在左侧 [上传] 或 [初始化] 表格！")
        return

    # 执行解析和追加
    try:
        info = extract_info_by_mode(text, mode)
        success, msg = append_data_to_workbook(st.session_state.workbook, info, mode)
        
        if success:
            st.session_state.status_msg = ("success", f"✅ {msg}")
            st.session_state.user_input = "" # 清空输入框
        else:
            st.session_state.status_msg = ("error", f"❌ {msg}")
    except Exception as e:
        st.session_state.status_msg = ("error", f"❌ 程序错误: {str(e)}")

# ================= 4. 页面布局 =================

st.title("📝 Excel 智能填表助手 (Web持久版)")

# --- Sidebar: 设置区 ---
with st.sidebar:
    st.header("1. 模式与文件")
    
    # 模式选择
    mode_options = ["福田统计", "爱心流动", "自定义"]
    selected_mode = st.radio("选择填表模式:", mode_options)
    st.session_state.current_mode = selected_mode # 更新状态

    st.markdown("---")
    
    # 文件操作类型
    file_op = st.radio("文件来源:", ["📂 上传现有 Excel", "✨ 新建空白表格"])
    
    if file_op == "📂 上传现有 Excel":
        uploaded_file = st.file_uploader("选择文件 (.xlsx)", type=["xlsx"])
        if uploaded_file:
            # 避免重复加载
            file_key = f"{uploaded_file.name}_{uploaded_file.size}"
            if st.session_state.last_loaded_key != file_key:
                try:
                    st.session_state.workbook = openpyxl.load_workbook(uploaded_file)
                    st.session_state.file_name = uploaded_file.name
                    st.session_state.last_loaded_key = file_key
                    st.success(f"已加载: {uploaded_file.name}")
                    st.rerun() # 重新运行以刷新预览
                except Exception as e:
                    st.error(f"加载失败: {e}")
    else:
        # 新建文件逻辑
        custom_headers = ""
        if selected_mode == "自定义":
            st.info("自定义模式下新建文件需指定列名")
            custom_headers = st.text_input("输入列名 (空格隔开)", value="姓名 电话 备注")
            
        if st.button("🚀 初始化新表格", type="primary"):
            st.session_state.workbook = create_blank_workbook(selected_mode, custom_headers)
            prefix = {"福田统计": "福田表", "爱心流动": "爱心表", "自定义": "自定义表"}
            st.session_state.file_name = f"{prefix[selected_mode]}_{datetime.now().strftime('%H%M')}.xlsx"
            st.session_state.last_loaded_key = f"NEW_{datetime.now().timestamp()}"
            st.success("新表格已创建！请在右侧开始录入。")
            st.rerun()

    st.markdown("---")
    st.caption("提示：所有操作都在内存中进行，离开页面前请务必点击右侧的【下载】按钮。")

# --- Main: 操作区 ---

col_input, col_preview = st.columns([1, 1.2])

# 左侧：输入
with col_input:
    st.subheader(f"2. 数据录入 ({st.session_state.current_mode})")
    
    placeholder_text = "姓名：张三\n电话：138000..."
    if st.session_state.current_mode == "爱心流动":
        placeholder_text = "被流动人：李四\n类型：爱心(1份)\n流动人：王五..."
        
    st.text_area(
        "在此粘贴文本:",
        height=300,
        key="user_input",
        placeholder=placeholder_text
    )
    
    # 提交按钮
    st.button("⚡ 解析并追加", type="primary", on_click=submit_data, use_container_width=True)
    
    # 消息反馈
    if st.session_state.status_msg:
        m_type, m_text = st.session_state.status_msg
        if m_type == "success": st.success(m_text)
        elif m_type == "error": st.error(m_text)
        elif m_type == "warning": st.warning(m_text)

# 右侧：预览与下载
with col_preview:
    st.subheader("3. 结果预览")
    
    if st.session_state.workbook:
        try:
            # 将 Workbook 转为 DataFrame 用于展示
            ws = st.session_state.workbook.active
            data = list(ws.values)
            
            if data:
                headers = data[0]
                rows = data[1:] if len(data) > 1 else []
                df = pd.DataFrame(rows, columns=headers)
                
                # 展示统计
                st.info(f"当前表格共有 **{len(rows)}** 条数据")
                
                # 可交互表格
                st.dataframe(df, use_container_width=True, height=350, hide_index=True)
                
                # 下载区
                st.markdown("### 📥 导出文件")
                excel_data = to_excel_bytes(st.session_state.workbook)
                
                col_d1, col_d2 = st.columns([3, 1])
                with col_d1:
                    new_name = st.text_input("文件名:", value=st.session_state.file_name, label_visibility="collapsed")
                with col_d2:
                    st.download_button(
                        label="下载",
                        data=excel_data,
                        file_name=new_name,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
            else:
                st.warning("表格是空的。")
        except Exception as e:
            st.error(f"预览生成错误: {e}")
    else:
        st.info("👈 请先在左侧侧边栏 [上传] 或 [新建] 表格")